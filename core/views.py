from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Produto, Receita, Ingrediente, Compra, ItemCompra, Cliente, Pedido, ItemPedido, ItemPedidoConsumo
import json
from datetime import date, timedelta, datetime
from collections import defaultdict
from django.core.paginator import Paginator
from decimal import Decimal, InvalidOperation
import calendar
import csv
from django.http import HttpResponse

STATUS_CHOICES = ['Pendente', 'Confirmado', 'Em produção', 'Pronto', 'Entregue', 'Cancelado']
FORMA_PAGAMENTO_CHOICES = ['PIX', 'Dinheiro', 'Cartão de crédito', 'Cartão de débito', 'Transferência', 'Outro']


def _calcular_custo_unitario_produto(produto):
    receitas = Receita.objects.filter(produto=produto).select_related('ingrediente')
    return sum((Decimal(r.quantidade) * Decimal(r.ingrediente.custo_unitario) for r in receitas), Decimal('0'))


def _registrar_snapshot_consumo_item(item_pedido):
    ItemPedidoConsumo.objects.filter(item_pedido=item_pedido).delete()
    receitas = Receita.objects.filter(produto=item_pedido.produto).select_related('ingrediente')
    for receita in receitas:
        ItemPedidoConsumo.objects.create(
            item_pedido=item_pedido,
            ingrediente=receita.ingrediente,
            ingrediente_nome=receita.ingrediente.nome,
            quantidade_por_unidade=receita.quantidade,
        )


def _normalizar_chave(texto):
    return str(texto or '').strip().lower().replace(' ', '_')


def _ler_linhas_planilha(arquivo):
    nome = arquivo.name.lower()

    if nome.endswith('.csv'):
        conteudo = arquivo.read().decode('utf-8-sig')
        leitor = csv.DictReader(conteudo.splitlines())
        return [dict(linha) for linha in leitor]

    if nome.endswith('.xlsx'):
        try:
            from openpyxl import load_workbook
        except Exception:
            raise RuntimeError('Para importar .xlsx, instale o pacote openpyxl no ambiente Python.')

        wb = load_workbook(filename=arquivo, data_only=True)
        ws = wb.active

        cabecalho = []
        linhas = []
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if idx == 1:
                cabecalho = [str(c or '').strip() for c in row]
                continue
            if not any(c is not None and str(c).strip() for c in row):
                continue
            registro = {}
            for i, coluna in enumerate(cabecalho):
                if not coluna:
                    continue
                registro[coluna] = row[i] if i < len(row) else None
            linhas.append(registro)
        return linhas

    raise RuntimeError('Formato não suportado. Use arquivos .csv ou .xlsx')


def importar_planilha(request):
    if request.method == 'POST':
        tipo = request.POST.get('tipo')
        arquivo = request.FILES.get('arquivo')

        if not arquivo:
            messages.error(request, 'Selecione um arquivo para importar.')
            return redirect('/importar/')

        try:
            linhas = _ler_linhas_planilha(arquivo)
        except RuntimeError as e:
            messages.error(request, str(e))
            return redirect('/importar/')
        except Exception:
            messages.error(request, 'Não foi possível ler a planilha. Verifique o arquivo.')
            return redirect('/importar/')

        criados = 0
        atualizados = 0
        ignorados = 0

        for linha in linhas:
            normalizada = {_normalizar_chave(k): v for k, v in linha.items()}

            if tipo == 'ingredientes':
                nome = normalizada.get('item') or normalizada.get('nome')
                valor_pacote = normalizada.get('valor_do_pacote') or normalizada.get('valor_pacote') or 0
                gramas = normalizada.get('gramas_por_unidade') or normalizada.get('gramas_unidade') or 0
                estoque_inicial = normalizada.get('estoque_inicial') or normalizada.get('estoque') or 0

                if not nome:
                    ignorados += 1
                    continue

                ingrediente, created = Ingrediente.objects.get_or_create(nome=str(nome).strip())
                ingrediente.ativo = True
                ingrediente.valor_pacote = valor_pacote or 0
                ingrediente.gramas_por_unidade = gramas or 0
                if estoque_inicial not in (None, ''):
                    ingrediente.estoque_atual = estoque_inicial
                ingrediente.save()

                if created:
                    criados += 1
                else:
                    atualizados += 1

            elif tipo == 'produtos':
                nome = normalizada.get('produto') or normalizada.get('nome')
                preco = normalizada.get('preco') or normalizada.get('preco_venda') or 0

                if not nome:
                    ignorados += 1
                    continue

                produto, created = Produto.objects.get_or_create(nome=str(nome).strip())
                produto.preco_venda = preco or 0
                produto.save()

                if created:
                    criados += 1
                else:
                    atualizados += 1

            else:
                messages.error(request, 'Tipo de importação inválido.')
                return redirect('/importar/')

        messages.success(
            request,
            f'Importação concluída: {criados} criados, {atualizados} atualizados, {ignorados} ignorados.'
        )
        return redirect('/importar/')

    return render(request, 'importar_planilha.html')


def _valor_planilha(valor):
    if isinstance(valor, Decimal):
        return float(valor)
    if isinstance(valor, datetime) and valor.tzinfo is not None:
        return valor.replace(tzinfo=None)
    return valor


def _estilizar_aba_planilha(ws, cabecalhos, tipos_coluna=None, destaque_linhas=None):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    tipos_coluna = tipos_coluna or {}
    destaque_linhas = destaque_linhas or {}
    header_fill = PatternFill(fill_type='solid', fgColor='7C3AED')
    header_font = Font(color='FFFFFF', bold=True)
    zebra_fill = PatternFill(fill_type='solid', fgColor='F5F3FF')
    entregue_fill = PatternFill(fill_type='solid', fgColor='DCFCE7')
    cancelado_fill = PatternFill(fill_type='solid', fgColor='FEE2E2')
    thin = Side(border_style='thin', color='DDD6FE')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    for col_idx, cabecalho in enumerate(cabecalhos, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    for row_idx in range(2, ws.max_row + 1):
        is_even = (row_idx % 2) == 0
        tipo_destaque = destaque_linhas.get(row_idx)
        for col_idx, cabecalho in enumerate(cabecalhos, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(vertical='center')

            tipo = tipos_coluna.get(cabecalho)
            if tipo == 'moeda':
                cell.number_format = 'R$ #,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif tipo == 'data':
                cell.number_format = 'DD/MM/YYYY'
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif tipo == 'data_hora':
                cell.number_format = 'DD/MM/YYYY HH:MM'
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif tipo == 'inteiro':
                cell.number_format = '0'
                cell.alignment = Alignment(horizontal='center', vertical='center')

            if tipo_destaque == 'entregue':
                cell.fill = entregue_fill
            elif tipo_destaque == 'cancelado':
                cell.fill = cancelado_fill
            elif is_even:
                cell.fill = zebra_fill

    for col_idx, cabecalho in enumerate(cabecalhos, start=1):
        largura = len(str(cabecalho)) + 4
        for row_idx in range(2, ws.max_row + 1):
            valor = ws.cell(row=row_idx, column=col_idx).value
            if valor is None:
                continue
            largura = max(largura, len(str(valor)) + 2)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(largura, 12), 48)


def _adicionar_aba_planilha(wb, nome_aba, cabecalhos, linhas, tipos_coluna=None, destaque_linhas=None):
    ws = wb.create_sheet(title=nome_aba)
    ws.append(cabecalhos)
    for linha in linhas:
        ws.append([_valor_planilha(v) for v in linha])
    _estilizar_aba_planilha(ws, cabecalhos, tipos_coluna, destaque_linhas)


def exportar_planilha(request):
    try:
        from openpyxl import Workbook
    except Exception:
        messages.error(request, 'Para exportar planilha .xlsx, instale o pacote openpyxl no ambiente Python.')
        return redirect('/')

    wb = Workbook()
    wb.remove(wb.active)

    compras = Compra.objects.all().order_by('-data', '-id')
    itens_pedido = ItemPedido.objects.select_related('pedido', 'produto').order_by('-pedido__data_entrega', '-pedido_id', 'id')
    pedidos = Pedido.objects.select_related('cliente').order_by('-data_entrega', '-id')

    pedidos_validos = [p for p in pedidos if p.status != 'Cancelado']
    faturamento_bruto = sum((p.total() for p in pedidos_validos), Decimal('0'))
    faturamento_liquido = sum((p.total_com_entrega() for p in pedidos_validos), Decimal('0'))
    custo_total_producao = sum(
        (item.custo_total_producao() for item in itens_pedido if item.pedido and item.pedido.status != 'Cancelado'),
        Decimal('0')
    )
    lucro_estimado = faturamento_liquido - custo_total_producao
    total_compras = sum((Decimal(c.total_final) for c in compras), Decimal('0'))

    _adicionar_aba_planilha(
        wb,
        'Resumo',
        ['Indicador', 'Quantidade', 'Valor (R$)'],
        [
            ['Pedidos totais', len(pedidos), None],
            ['Pedidos entregues', len([p for p in pedidos if p.status == 'Entregue']), None],
            ['Pedidos cancelados', len([p for p in pedidos if p.status == 'Cancelado']), None],
            ['Compras registradas', len(compras), None],
            ['Faturamento bruto pedidos validos', None, faturamento_bruto],
            ['Faturamento com entrega pedidos validos', None, faturamento_liquido],
            ['Custo total de producao pedidos validos', None, custo_total_producao],
            ['Total de compras', None, total_compras],
            ['Lucro estimado', None, lucro_estimado],
        ],
        {
            'Quantidade': 'inteiro',
            'Valor (R$)': 'moeda',
        },
    )

    ingredientes = Ingrediente.objects.all().order_by('nome')
    _adicionar_aba_planilha(
        wb,
        'Ingredientes',
        ['Ingrediente', 'Ativo', 'Valor pacote (R$)', 'Gramas por unidade', 'Custo por grama (R$)', 'Estoque atual (pacotes)'],
        [
            [
                i.nome,
                'Sim' if i.ativo else 'Nao',
                i.valor_pacote,
                i.gramas_por_unidade,
                i.custo_unitario,
                i.estoque_atual,
            ]
            for i in ingredientes
        ],
        {
            'Valor pacote (R$)': 'moeda',
            'Custo por grama (R$)': 'moeda',
        },
    )

    produtos = Produto.objects.all().order_by('nome')
    _adicionar_aba_planilha(
        wb,
        'Produtos',
        ['Produto', 'Preco de venda (R$)', 'Imagem'],
        [
            [p.nome, p.preco_venda, p.imagem.name if p.imagem else '']
            for p in produtos
        ],
        {
            'Preco de venda (R$)': 'moeda',
        },
    )

    receitas = Receita.objects.select_related('produto', 'ingrediente').order_by('produto__nome', 'ingrediente__nome')
    _adicionar_aba_planilha(
        wb,
        'Receitas',
        ['Produto', 'Ingrediente', 'Quantidade (gramas)', 'Custo estimado (R$)'],
        [
            [
                r.produto.nome,
                r.ingrediente.nome,
                r.quantidade,
                r.custo(),
            ]
            for r in receitas
        ],
        {
            'Custo estimado (R$)': 'moeda',
        },
    )

    _adicionar_aba_planilha(
        wb,
        'Compras',
        ['Data', 'Local', 'Desconto (R$)', 'Frete (R$)', 'Total ingredientes (R$)', 'Total calculado (R$)', 'Total final (R$)'],
        [
            [
                c.data,
                c.local,
                c.desconto,
                c.frete,
                c.total_ingredientes,
                c.total_calculado,
                c.total_final,
            ]
            for c in compras
        ],
        {
            'Data': 'data',
            'Desconto (R$)': 'moeda',
            'Frete (R$)': 'moeda',
            'Total ingredientes (R$)': 'moeda',
            'Total calculado (R$)': 'moeda',
            'Total final (R$)': 'moeda',
        },
    )

    itens_compra = ItemCompra.objects.select_related('compra', 'ingrediente').order_by('-compra__data', '-compra_id', 'id')
    _adicionar_aba_planilha(
        wb,
        'ItensCompra',
        [
            'Data da compra',
            'Local da compra',
            'Ingrediente',
            'Quantidade (gramas)',
            'Quantidade (pacotes)',
            'Custo unitario (R$)',
            'Custo total (R$)',
        ],
        [
            [
                item.compra.data,
                item.compra.local,
                item.ingrediente.nome,
                item.quantidade,
                item.quantidade_pacotes(),
                item.custo_unitario_compra,
                item.custo_total,
            ]
            for item in itens_compra
        ],
        {
            'Data da compra': 'data',
            'Custo unitario (R$)': 'moeda',
            'Custo total (R$)': 'moeda',
        },
    )

    clientes = Cliente.objects.all().order_by('nome')
    _adicionar_aba_planilha(
        wb,
        'Clientes',
        ['Cliente', 'Telefone', 'Endereco'],
        [[c.nome, c.telefone, c.endereco] for c in clientes],
    )

    linhas_pedidos = [
        [
            p.id,
            p.cliente.nome,
            p.cliente.telefone,
            p.data_pedido,
            p.data_entrega,
            p.taxa_entrega,
            p.desconto,
            p.forma_pagamento,
            p.status,
            'Sim' if p.baixa_estoque_aprovada else 'Nao',
            p.total(),
            p.total_com_entrega(),
            p.observacoes,
        ]
        for p in pedidos
    ]
    destaque_linhas_pedidos = {}
    for idx, p in enumerate(pedidos, start=2):
        if p.status == 'Entregue':
            destaque_linhas_pedidos[idx] = 'entregue'
        elif p.status == 'Cancelado':
            destaque_linhas_pedidos[idx] = 'cancelado'

    _adicionar_aba_planilha(
        wb,
        'Pedidos',
        [
            'Pedido ID',
            'Cliente',
            'Telefone',
            'Data do pedido',
            'Data de entrega',
            'Taxa de entrega (R$)',
            'Desconto (R$)',
            'Forma de pagamento',
            'Status',
            'Baixa de estoque aprovada',
            'Total (R$)',
            'Total com entrega (R$)',
            'Observacoes',
        ],
        linhas_pedidos,
        {
            'Pedido ID': 'inteiro',
            'Data do pedido': 'data_hora',
            'Data de entrega': 'data',
            'Taxa de entrega (R$)': 'moeda',
            'Desconto (R$)': 'moeda',
            'Total (R$)': 'moeda',
            'Total com entrega (R$)': 'moeda',
        },
        destaque_linhas_pedidos,
    )

    _adicionar_aba_planilha(
        wb,
        'ItensPedido',
        [
            'Data de entrega',
            'Pedido ID',
            'Produto',
            'Quantidade',
            'Preco unitario (R$)',
            'Subtotal (R$)',
            'Custo unitario producao (R$)',
            'Custo total producao (R$)',
        ],
        [
            [
                item.pedido.data_entrega,
                item.pedido_id,
                item.nome_produto(),
                item.quantidade,
                item.preco_unitario,
                item.subtotal(),
                item.custo_unitario_producao,
                item.custo_total_producao(),
            ]
            for item in itens_pedido
        ],
        {
            'Data de entrega': 'data',
            'Pedido ID': 'inteiro',
            'Preco unitario (R$)': 'moeda',
            'Subtotal (R$)': 'moeda',
            'Custo unitario producao (R$)': 'moeda',
            'Custo total producao (R$)': 'moeda',
        },
    )

    nome_arquivo = f"dados_doces_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
    wb.save(response)
    return response


def calcular_necessidades_pedido(pedido):
    totais = defaultdict(lambda: Decimal('0'))
    itens = ItemPedido.objects.filter(pedido=pedido).select_related('produto').prefetch_related('itempedidoconsumo_set')
    for item in itens:
        consumos = list(item.itempedidoconsumo_set.all())
        if consumos:
            for consumo in consumos:
                if consumo.ingrediente_id:
                    totais[consumo.ingrediente_id] += Decimal(consumo.quantidade_por_unidade) * Decimal(item.quantidade)
            continue

        # Fallback para itens legados sem snapshot.
        receitas = Receita.objects.filter(produto=item.produto).select_related('ingrediente')
        for r in receitas:
            totais[r.ingrediente_id] += Decimal(r.quantidade) * Decimal(item.quantidade)

    ingredientes = Ingrediente.objects.filter(id__in=totais.keys())
    por_id = {i.id: i for i in ingredientes}

    preview = []
    for ingrediente_id, necessario in totais.items():
        ingrediente = por_id.get(ingrediente_id)
        if not ingrediente:
            continue
        estoque_atual = Decimal(ingrediente.estoque_atual)
        gramas_unidade = Decimal(ingrediente.gramas_por_unidade or 0)
        if gramas_unidade > 0:
            necessario_pacotes = necessario / gramas_unidade
            saldo_final = estoque_atual - necessario_pacotes
            suficiente = saldo_final >= 0
        else:
            necessario_pacotes = Decimal('0')
            saldo_final = estoque_atual
            suficiente = False

        preview.append({
            'ingrediente': ingrediente,
            'necessario': necessario,
            'necessario_pacotes': necessario_pacotes,
            'estoque_atual': estoque_atual,
            'saldo_final': saldo_final,
            'suficiente': suficiente,
        })
    preview.sort(key=lambda x: x['ingrediente'].nome.lower())
    return preview


def dashboard(request):
    hoje = date.today()
    pedidos_hoje = Pedido.objects.filter(data_entrega=hoje).count()
    pedidos_pendentes = Pedido.objects.filter(status__in=['Pendente', 'Confirmado', 'Em produção']).count()
    estoque_baixo = Ingrediente.objects.filter(ativo=True, estoque_atual__lte=1).order_by('estoque_atual', 'nome')
    return render(request, "dashboard.html", {
        'pedidos_hoje': pedidos_hoje,
        'pedidos_pendentes': pedidos_pendentes,
        'estoque_baixo': estoque_baixo,
        'estoque_baixo_total': estoque_baixo.count(),
    })


def ingredientes(request):
    if request.method == "POST":
        action = request.POST.get("action", "create")

        if action == 'edit_price':
            ingrediente_id = request.POST.get('ingrediente_id')
            ingrediente = Ingrediente.objects.filter(id=ingrediente_id, ativo=True).first()
            if not ingrediente:
                messages.error(request, 'Ingrediente não encontrado para edição.')
                return redirect('/ingredientes/')

            ingrediente.valor_pacote = request.POST.get('valor_pacote') or ingrediente.valor_pacote
            ingrediente.save()
            messages.success(request, f'Preço de "{ingrediente.nome}" atualizado com sucesso.')
            return redirect('/ingredientes/')

        estoque_inicial = request.POST.get("estoque_inicial") or 0
        Ingrediente.objects.create(
            nome=request.POST["nome"],
            ativo=True,
            valor_pacote=request.POST["valor_pacote"],
            gramas_por_unidade=request.POST["gramas_por_unidade"],
            estoque_atual=estoque_inicial,
        )
        messages.success(request, 'Ingrediente cadastrado com sucesso.')
        return redirect('/ingredientes/')

    lista = Ingrediente.objects.filter(ativo=True).order_by('nome')
    return render(request, "ingredientes.html", {"ingredientes": lista})


def ingrediente_excluir(request, ingrediente_id):
    if request.method != "POST":
        return redirect('/ingredientes/')

    ingrediente = Ingrediente.objects.filter(id=ingrediente_id, ativo=True).first()
    if not ingrediente:
        messages.error(request, 'Ingrediente não encontrado.')
        return redirect('/ingredientes/')

    nome = ingrediente.nome
    ingrediente.ativo = False
    ingrediente.save(update_fields=['ativo'])
    messages.success(request, f'Ingrediente "{nome}" removido do cadastro sem apagar o histórico.')
    return redirect('/ingredientes/')


def compra_excluir(request, compra_id):
    if request.method != "POST":
        return redirect('/compras/')

    compra = Compra.objects.filter(id=compra_id).prefetch_related('itemcompra_set__ingrediente').first()
    if not compra:
        messages.error(request, 'Compra não encontrada.')
        return redirect('/compras/')

    for item in compra.itemcompra_set.all():
        ingrediente = item.ingrediente
        if not ingrediente:
            continue

        pacotes = Decimal(item.quantidade_pacotes_registrada or 0)
        if pacotes <= 0:
            gramas_unidade = Decimal(ingrediente.gramas_por_unidade or 0)
            if gramas_unidade > 0:
                pacotes = Decimal(item.quantidade) / gramas_unidade

        ingrediente.estoque_atual = Decimal(ingrediente.estoque_atual) - pacotes
        ingrediente.save()

    compra.delete()
    messages.success(request, 'Compra excluída com sucesso.')
    return redirect('/compras/')


# 🍫 NOVO PRODUTO + RECEITA INLINE

def produto_novo(request):
    if request.method == "POST":
        produto = Produto.objects.create(
            nome=request.POST["nome"],
            preco_venda=request.POST["preco"],
            imagem=request.FILES.get("imagem")
        )
        for ingr in Ingrediente.objects.filter(ativo=True):
            val = request.POST.get(f"qtd_{ingr.id}", "").strip()
            if val:
                try:
                    qtd = float(val)
                    if qtd > 0:
                        Receita.objects.create(produto=produto, ingrediente=ingr, quantidade=qtd)
                except ValueError:
                    pass
        return redirect('/produtos/')
    ingredientes_q = request.GET.get("q", "").strip()
    ingredientes_qs = Ingrediente.objects.filter(ativo=True).order_by("nome")
    if ingredientes_q:
        ingredientes_qs = ingredientes_qs.filter(nome__icontains=ingredientes_q)
    paginator = Paginator(ingredientes_qs, 12)
    page_obj_ingredientes = paginator.get_page(request.GET.get("page"))
    return render(request, "produto_novo.html", {
        "ingredientes": page_obj_ingredientes,
        "ingredientes_q": ingredientes_q,
    })


# 👁️ LISTA DE PRODUTOS

def produtos(request):
    lista = Produto.objects.all().order_by('nome')
    produtos_data = []

    for produto in lista:
        receitas = Receita.objects.filter(produto=produto).select_related('ingrediente')
        custo_receita = sum(r.custo() for r in receitas)
        margem_valor = produto.preco_venda - custo_receita
        if custo_receita > 0:
            margem_percentual = (margem_valor / custo_receita) * 100
        else:
            margem_percentual = None

        produtos_data.append({
            'produto': produto,
            'custo_receita': custo_receita,
            'margem_valor': margem_valor,
            'margem_percentual': margem_percentual,
        })

    return render(request, "produtos.html", {"produtos_data": produtos_data})


def produto_editar(request, produto_id):
    produto = Produto.objects.filter(id=produto_id).first()
    if not produto:
        messages.error(request, 'Produto não encontrado.')
        return redirect('/produtos/')

    if request.method == 'POST':
        produto.nome = request.POST.get('nome', produto.nome)
        produto.preco_venda = request.POST.get('preco', produto.preco_venda)

        if request.POST.get('remover_imagem') == 'on' and produto.imagem:
            produto.imagem.delete(save=False)
            produto.imagem = None

        nova_imagem = request.FILES.get('imagem')
        if nova_imagem:
            if produto.imagem:
                produto.imagem.delete(save=False)
            produto.imagem = nova_imagem

        produto.save()
        messages.success(request, 'Produto atualizado com sucesso.')
        return redirect('/produtos/')

    return render(request, 'produto_editar.html', {'produto': produto})


def produto_excluir(request, produto_id):
    if request.method != "POST":
        return redirect('/produtos/')

    produto = Produto.objects.filter(id=produto_id).first()
    if not produto:
        messages.error(request, 'Produto não encontrado.')
        return redirect('/produtos/')

    itens_vinculados = ItemPedido.objects.filter(produto=produto).select_related('pedido')
    itens_bloqueantes = itens_vinculados.exclude(pedido__status__in=['Entregue', 'Cancelado'])
    if itens_bloqueantes.exists():
        messages.error(
            request,
            'Não é possível excluir: este produto possui pedidos em andamento (apenas Entregue/Cancelado permitem exclusão).'
        )
        return redirect('/produtos/')

    if itens_vinculados.exists():
        itens_vinculados.update(produto_nome=produto.nome, produto=None)

    nome = produto.nome
    produto.delete()
    messages.success(request, f'Produto "{nome}" removido com sucesso.')
    return redirect('/produtos/')


# 📋 RECEITA DE UM PRODUTO

def receita_produto(request, produto_id):
    produto = Produto.objects.get(id=produto_id)
    if request.method == "POST":
        action = request.POST.get('action', 'add')
        if action == 'remove':
            receita = Receita.objects.filter(id=request.POST.get('receita_id'), produto=produto).first()
            if not receita:
                messages.error(request, 'Ingrediente da receita não encontrado.')
                return redirect(f'/produtos/{produto_id}/receita')
            receita.delete()
            messages.success(request, 'Ingrediente removido da receita com sucesso.')
            return redirect(f'/produtos/{produto_id}/receita')

        Receita.objects.create(
            produto=produto,
            ingrediente_id=request.POST["ingrediente"],
            quantidade=request.POST["quantidade"]
        )
        messages.success(request, 'Ingrediente adicionado na receita.')
        return redirect(f'/produtos/{produto_id}/receita')

    ingrediente_q = request.GET.get("ingrediente_q", "").strip()
    receitas = Receita.objects.filter(produto=produto).select_related("ingrediente")
    if ingrediente_q:
        receitas = receitas.filter(ingrediente__nome__icontains=ingrediente_q)

    paginator = Paginator(receitas.order_by("ingrediente__nome"), 10)
    page_obj = paginator.get_page(request.GET.get("page"))

    ingredientes = Ingrediente.objects.filter(ativo=True).order_by('nome')
    custo_total = sum(r.custo() for r in Receita.objects.filter(produto=produto).select_related("ingrediente"))
    return render(request, "receita.html", {
        "produto": produto,
        "receitas": page_obj,
        "ingredientes": ingredientes,
        "custo_total": custo_total,
        "ingrediente_q": ingrediente_q,
    })


# 📋 COMPRAS - ESTOQUE

def compras(request):
    if request.method == "POST":
        data_str = request.POST.get("data") or str(date.today())
        local = request.POST.get("local", "")
        try:
            desconto = Decimal(str(request.POST.get("desconto") or 0))
        except (InvalidOperation, TypeError, ValueError):
            desconto = Decimal('0')
        try:
            frete = Decimal(str(request.POST.get("frete") or 0))
        except (InvalidOperation, TypeError, ValueError):
            frete = Decimal('0')
        total_final_informado = request.POST.get("total_final", "").strip()
        compra = Compra.objects.create(
            data=data_str,
            local=local,
            desconto=desconto,
            frete=frete,
        )

        ingrediente_ids = request.POST.getlist("ingrediente_id")
        quantidades_pacotes = request.POST.getlist("quantidade_pacotes")
        custos_unitarios = request.POST.getlist("custo_unitario")

        total_ingredientes = Decimal('0')
        for i, ingr_id in enumerate(ingrediente_ids):
            if not ingr_id:
                continue
            qtd_p = quantidades_pacotes[i] if i < len(quantidades_pacotes) else "0"
            cst_unit = custos_unitarios[i] if i < len(custos_unitarios) else "0"

            try:
                ingrediente = Ingrediente.objects.filter(id=ingr_id).first()
                if not ingrediente:
                    continue

                gramas_unidade = Decimal(ingrediente.gramas_por_unidade or 0)
                qtd_pacotes_dec = Decimal(str(qtd_p or 0))
                custo_unitario_compra = Decimal(str(cst_unit or 0))

                if qtd_pacotes_dec > 0 and custo_unitario_compra >= 0:
                    qtd_gramas_dec = qtd_pacotes_dec * gramas_unidade if gramas_unidade > 0 else Decimal('0')
                    custo_total_item = qtd_pacotes_dec * custo_unitario_compra

                    item = ItemCompra.objects.create(
                        compra=compra,
                        ingrediente=ingrediente,
                        quantidade=qtd_gramas_dec,
                        custo_unitario_compra=custo_unitario_compra,
                        custo_total=custo_total_item,
                        quantidade_pacotes_registrada=qtd_pacotes_dec,
                    )
                    total_ingredientes += custo_total_item

                    if gramas_unidade > 0:
                        pacotes_entrada = Decimal(item.quantidade) / gramas_unidade
                        ingrediente.estoque_atual = Decimal(ingrediente.estoque_atual) + pacotes_entrada
                    else:
                        ingrediente.estoque_atual = Decimal(ingrediente.estoque_atual) + qtd_pacotes_dec

                    # A última compra define o valor do pacote para custos futuros.
                    ingrediente.valor_pacote = custo_unitario_compra
                    ingrediente.save()
            except ValueError:
                pass
            except Exception:
                pass

        total_calculado = total_ingredientes + frete - desconto
        if total_calculado < 0:
            total_calculado = Decimal('0')

        try:
            total_final = Decimal(str(total_final_informado)) if total_final_informado else total_calculado
        except (InvalidOperation, TypeError, ValueError):
            total_final = total_calculado

        compra.total_ingredientes = total_ingredientes
        compra.total_calculado = total_calculado
        compra.total_final = total_final
        compra.save(update_fields=['total_ingredientes', 'total_calculado', 'total_final'])

    ingredientes = Ingrediente.objects.filter(ativo=True).order_by('nome')
    produtos = Produto.objects.all()

    receitas_por_produto = {}
    for produto in produtos:
        receitas = Receita.objects.filter(produto=produto).select_related('ingrediente')
        receitas_por_produto[str(produto.id)] = [
            {
                'id': r.ingrediente.id,
                'nome': r.ingrediente.nome,
                'pacotes_sugeridos': float(
                    Decimal(r.quantidade) / Decimal(r.ingrediente.gramas_por_unidade)
                ) if Decimal(r.ingrediente.gramas_por_unidade or 0) > 0 else 0
            }
            for r in receitas
        ]

    compras_list = Compra.objects.prefetch_related('itemcompra_set__ingrediente').order_by("-data")
    return render(request, "compras.html", {
        "ingredientes": ingredientes,
        "produtos": produtos,
        "compras": compras_list,
        "receitas_json": json.dumps(receitas_por_produto),
        "today": date.today(),
    })


# 📦 PEDIDOS

def pedidos(request):
    status_filter = request.GET.get('status', '')
    qs = Pedido.objects.select_related('cliente').prefetch_related('itempedido_set__produto')
    if status_filter:
        qs = qs.filter(status=status_filter)
    lista = qs.order_by('data_entrega')
    return render(request, 'pedidos.html', {
        'pedidos': lista,
        'status_filter': status_filter,
        'status_choices': STATUS_CHOICES,
    })


def pedido_novo(request):
    if request.method == 'POST':
        telefone = request.POST.get('cliente_telefone', '').strip()
        nome = request.POST.get('cliente_nome', '').strip()
        endereco = request.POST.get('cliente_endereco', '').strip()

        cliente, _ = Cliente.objects.get_or_create(
            telefone=telefone,
            defaults={'nome': nome, 'endereco': endereco}
        )
        if nome:
            cliente.nome = nome
        if endereco:
            cliente.endereco = endereco
        cliente.save()

        pedido = Pedido.objects.create(
            cliente=cliente,
            data_entrega=request.POST['data_entrega'],
            taxa_entrega=request.POST.get('taxa_entrega') or 0,
            desconto=request.POST.get('desconto') or 0,
            forma_pagamento=request.POST.get('forma_pagamento') or '',
            observacoes=request.POST.get('observacoes', ''),
            status='Pendente'
        )

        produto_ids = request.POST.getlist('produto_id')
        quantidades = request.POST.getlist('quantidade')
        for i, pid in enumerate(produto_ids):
            if not pid:
                continue
            try:
                qtd = int(quantidades[i]) if i < len(quantidades) else 1
                if qtd > 0:
                    produto = Produto.objects.get(id=pid)
                    custo_unitario_producao = _calcular_custo_unitario_produto(produto)
                    item_pedido = ItemPedido.objects.create(
                        pedido=pedido,
                        produto=produto,
                        produto_nome=produto.nome,
                        quantidade=qtd,
                        preco_unitario=produto.preco_venda,
                        custo_unitario_producao=custo_unitario_producao,
                    )
                    _registrar_snapshot_consumo_item(item_pedido)
            except (ValueError, Produto.DoesNotExist):
                pass

        return redirect(f'/pedidos/{pedido.id}/')

    produtos_list = Produto.objects.all()
    clientes = Cliente.objects.order_by('nome')
    return render(request, 'pedido_novo.html', {
        'produtos': produtos_list,
        'clientes': clientes,
        'today': date.today(),
        'status_choices': STATUS_CHOICES,
        'forma_pagamento_choices': FORMA_PAGAMENTO_CHOICES,
    })


def pedido_detalhe(request, pedido_id):
    pedido = Pedido.objects.select_related('cliente').get(id=pedido_id)

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'status':
            pedido.status = request.POST['status']
            pedido.save()
        elif action == 'financeiro':
            pedido.taxa_entrega = request.POST.get('taxa_entrega') or 0
            pedido.desconto = request.POST.get('desconto') or 0
            pedido.forma_pagamento = request.POST.get('forma_pagamento') or ''
            pedido.save()
            messages.success(request, 'Dados financeiros atualizados.')
        elif action == 'add_item':
            if pedido.baixa_estoque_aprovada:
                messages.error(request, 'Baixa de estoque já aprovada. Não é possível alterar itens deste pedido.')
                return redirect(f'/pedidos/{pedido_id}/')
            try:
                produto = Produto.objects.get(id=request.POST['produto_id'])
                custo_unitario_producao = _calcular_custo_unitario_produto(produto)
                item_pedido = ItemPedido.objects.create(
                    pedido=pedido,
                    produto=produto,
                    produto_nome=produto.nome,
                    quantidade=int(request.POST['quantidade']),
                    preco_unitario=produto.preco_venda,
                    custo_unitario_producao=custo_unitario_producao,
                )
                _registrar_snapshot_consumo_item(item_pedido)
            except (Produto.DoesNotExist, ValueError):
                pass
        elif action == 'remove_item':
            if pedido.baixa_estoque_aprovada:
                messages.error(request, 'Baixa de estoque já aprovada. Não é possível alterar itens deste pedido.')
                return redirect(f'/pedidos/{pedido_id}/')
            ItemPedido.objects.filter(id=request.POST['item_id'], pedido=pedido).delete()
        elif action == 'aprovar_baixa_estoque':
            if pedido.baixa_estoque_aprovada:
                messages.error(request, 'A baixa deste pedido já foi aprovada anteriormente.')
                return redirect(f'/pedidos/{pedido_id}/')

            preview = calcular_necessidades_pedido(pedido)
            if any(not p['suficiente'] for p in preview):
                messages.error(request, 'Estoque insuficiente para aprovar a baixa. Ajuste o estoque primeiro.')
                return redirect(f'/pedidos/{pedido_id}/')

            for p in preview:
                ingrediente = p['ingrediente']
                ingrediente.estoque_atual = p['saldo_final']
                ingrediente.save()

            pedido.baixa_estoque_aprovada = True
            pedido.save()
            messages.success(request, 'Baixa de estoque aprovada e aplicada com sucesso.')
        return redirect(f'/pedidos/{pedido_id}/')

    itens = ItemPedido.objects.filter(pedido=pedido).select_related('produto')
    produtos_list = Produto.objects.all()
    total = sum(i.subtotal() for i in itens)
    total_com_entrega = total + pedido.taxa_entrega
    estoque_preview = calcular_necessidades_pedido(pedido)
    faltas_estoque = [p for p in estoque_preview if not p['suficiente']]
    return render(request, 'pedido_detalhe.html', {
        'pedido': pedido,
        'itens': itens,
        'produtos': produtos_list,
        'total': total,
        'total_com_entrega': total_com_entrega,
        'status_choices': STATUS_CHOICES,
        'forma_pagamento_choices': FORMA_PAGAMENTO_CHOICES,
        'estoque_preview': estoque_preview,
        'tem_falta_estoque': len(faltas_estoque) > 0,
    })


def pedido_excluir(request, pedido_id):
    if request.method != "POST":
        return redirect('/pedidos/')

    pedido = Pedido.objects.filter(id=pedido_id).first()
    if not pedido:
        messages.error(request, 'Pedido não encontrado.')
        return redirect('/pedidos/')

    if pedido.baixa_estoque_aprovada:
        preview = calcular_necessidades_pedido(pedido)
        for p in preview:
            ingrediente = p['ingrediente']
            ingrediente.estoque_atual = Decimal(ingrediente.estoque_atual) + Decimal(p['necessario_pacotes'])
            ingrediente.save()

    pedido.delete()
    messages.success(request, 'Pedido excluído com sucesso.')
    return redirect('/pedidos/')


# 📅 AGENDA

def agenda(request):
    hoje = date.today()
    mes_param = request.GET.get('mes', '')
    if mes_param:
        try:
            ano, mes = map(int, mes_param.split('-'))
        except ValueError:
            ano, mes = hoje.year, hoje.month
    else:
        ano, mes = hoje.year, hoje.month

    primeiro_dia = date(ano, mes, 1)
    if mes == 12:
        prox_mes = date(ano + 1, 1, 1)
    else:
        prox_mes = date(ano, mes + 1, 1)

    pedidos_qs = Pedido.objects.filter(
        data_entrega__gte=primeiro_dia,
        data_entrega__lt=prox_mes
    ).select_related('cliente').prefetch_related('itempedido_set__produto').order_by('data_entrega')

    por_data = defaultdict(list)
    for p in pedidos_qs:
        por_data[p.data_entrega].append(p)

    cal = calendar.Calendar(firstweekday=6)
    semanas = []
    for semana in cal.monthdatescalendar(ano, mes):
        semana_items = []
        for d in semana:
            semana_items.append({
                'data': d,
                'no_mes': d.month == mes,
                'hoje': d == hoje,
                'pedidos': por_data.get(d, []),
            })
        semanas.append(semana_items)

    mes_anterior = primeiro_dia - timedelta(days=1)
    mes_seguinte = prox_mes

    return render(request, 'agenda.html', {
        'semanas': semanas,
        'mes_data': primeiro_dia,
        'mes_anterior': f"{mes_anterior.year}-{mes_anterior.month:02d}",
        'mes_seguinte': f"{mes_seguinte.year}-{mes_seguinte.month:02d}",
    })


# 🌐 PÁGINAS PÚBLICAS

def catalogo(request):
    produtos_list = Produto.objects.all()
    return render(request, 'catalogo.html', {'produtos': produtos_list})


def consulta_pedido(request):
    pedido = None
    erro = None
    if request.method == 'POST':
        pedido_id = request.POST.get('pedido_id', '').strip()
        telefone = request.POST.get('telefone', '').strip()
        if pedido_id and telefone:
            try:
                pedido = Pedido.objects.select_related('cliente').prefetch_related(
                    'itempedido_set__produto'
                ).get(id=pedido_id, cliente__telefone=telefone)
            except (Pedido.DoesNotExist, ValueError):
                erro = 'Pedido não encontrado. Verifique o número do pedido e o telefone cadastrado.'
        else:
            erro = 'Preencha o número do pedido e o telefone.'
    return render(request, 'consulta_pedido.html', {'pedido': pedido, 'erro': erro})
